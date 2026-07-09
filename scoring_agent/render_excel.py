"""Excel renderer.

Sheet 1 "Scorecard": dimensions x vendors matrix (native scores), weights,
normalized totals, native totals, and rank; cells colored by provenance
(extracted vs generated vs gate-fail) with per-cell rationale as comments.
Sheet 2 "Evidence & Notes": one row per scored cell with rationale + citations.
Sheet 3 "Scheme": the parsed native scheme (method, scale, weights, gates).
"""

from __future__ import annotations

from typing import Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Provenance, ScorecardResult, ScoreCell

# provenance -> fill color
_FILL = {
    Provenance.EXTRACTED: "D6EAD3",   # green  = real (extracted)
    Provenance.GENERATED: "DCE6F1",   # blue   = predicted (generated)
    Provenance.GATE_FAIL: "F4CCCC",   # red    = gate fail
    Provenance.NOT_SCORED: "F2F2F2",  # grey   = not scored
}
_IBM_BLUE = "0F62FE"
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="IBM Plex Sans", size=11)
_LABEL_FONT = Font(bold=True, name="IBM Plex Sans", size=10)
_BODY_FONT = Font(name="IBM Plex Sans", size=10)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center")


def _cell_display(cell: Optional[ScoreCell]) -> str:
    if cell is None:
        return ""
    if cell.provenance == Provenance.GATE_FAIL:
        return "FAIL"
    if cell.native_value:
        return cell.native_value
    if cell.value is not None:
        return f"{cell.value:g}"
    return ""


def _style_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _HEADER_FONT
    c.fill = PatternFill("solid", fgColor=_IBM_BLUE)
    c.alignment = _CENTER
    c.border = _BORDER
    return c


def render_excel(result: ScorecardResult, path: str) -> str:
    scheme = result.scheme
    weights = scheme.effective_weights()
    vendors = result.ranked_vendors()

    wb = Workbook()

    # ---- Sheet 1: Scorecard --------------------------------------------- #
    ws = wb.active
    ws.title = "Scorecard"
    ws.sheet_view.showGridLines = False

    _style_header(ws, 1, 1, "Evaluation Dimension")
    _style_header(ws, 1, 2, "Weight")
    for j, v in enumerate(vendors):
        _style_header(ws, 1, 3 + j, v.vendor)

    r = 2
    for dim in scheme.dimensions:
        lc = ws.cell(row=r, column=1, value=dim.name)
        lc.font = _LABEL_FONT
        lc.alignment = _WRAP
        lc.border = _BORDER
        wc = ws.cell(row=r, column=2, value=round(weights.get(dim.id, 0.0), 4))
        wc.number_format = "0.0%"
        wc.font = _BODY_FONT
        wc.alignment = _CENTER
        wc.border = _BORDER
        for j, v in enumerate(vendors):
            cell = v.cells.get(dim.id)
            gc = ws.cell(row=r, column=3 + j, value=_cell_display(cell))
            gc.font = _BODY_FONT
            gc.alignment = _CENTER
            gc.border = _BORDER
            prov = cell.provenance if cell else Provenance.NOT_SCORED
            gc.fill = PatternFill("solid", fgColor=_FILL.get(prov, "FFFFFF"))
            if cell and (cell.rationale or cell.evidence):
                note = cell.rationale or ""
                if cell.evidence:
                    note += "\n\nEvidence:\n" + "\n".join(
                        f"- {e.doc}"
                        + (f" p.{e.page}" if e.page else "")
                        + (f": “{e.quote}”" if e.quote else "")
                        for e in cell.evidence[:4]
                    )
                gc.comment = Comment(note[:2000], "scoring_agent")
        r += 1

    # summary rows
    def _summary_row(label, getter, number_format=None):
        nonlocal r
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _LABEL_FONT
        lc.border = _BORDER
        ws.cell(row=r, column=2).border = _BORDER
        for j, v in enumerate(vendors):
            val = getter(v)
            c = ws.cell(row=r, column=3 + j, value=val)
            c.font = _LABEL_FONT
            c.alignment = _CENTER
            c.border = _BORDER
            if number_format and isinstance(val, (int, float)):
                c.number_format = number_format
        r += 1

    r += 1
    _summary_row(
        "Normalized Total",
        lambda v: (round(v.normalized_total_pct, 1) / 100 if v.normalized_total_pct is not None else None),
        number_format="0.0%",
    )
    if any(v.native_total is not None for v in vendors):
        total_max = scheme.total_max_points
        label = "Native Total" + (f" (/{total_max:g})" if total_max else "")
        _summary_row(
            label,
            lambda v: (round(v.native_total, 1) if v.native_total is not None else None),
            number_format="0.0",
        )
    _summary_row(
        "Rank",
        lambda v: ("DQ" if v.disqualified else (v.rank if v.rank is not None else "")),
    )

    # column widths
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 9
    for j in range(len(vendors)):
        ws.column_dimensions[get_column_letter(3 + j)].width = 16

    # legend
    r += 1
    ws.cell(row=r, column=1, value="Legend:").font = _LABEL_FONT
    legend = [
        ("Extracted (real scoresheet)", Provenance.EXTRACTED),
        ("Generated (predicted)", Provenance.GENERATED),
        ("Minimum-requirement fail", Provenance.GATE_FAIL),
    ]
    for k, (txt, prov) in enumerate(legend):
        c = ws.cell(row=r, column=2 + k, value=txt)
        c.font = _BODY_FONT
        c.fill = PatternFill("solid", fgColor=_FILL[prov])
        c.alignment = _CENTER

    # ---- Sheet 2: Evidence & Notes -------------------------------------- #
    ev = wb.create_sheet("Evidence & Notes")
    headers = ["Vendor", "Dimension", "Score", "Provenance", "Confidence", "Rationale", "Evidence"]
    for i, h in enumerate(headers, start=1):
        _style_header(ev, 1, i, h)
    er = 2
    for v in vendors:
        for dim in scheme.dimensions:
            cell = v.cells.get(dim.id)
            if cell is None:
                continue
            ev.cell(row=er, column=1, value=v.vendor).font = _BODY_FONT
            ev.cell(row=er, column=2, value=dim.name).font = _BODY_FONT
            ev.cell(row=er, column=3, value=_cell_display(cell)).font = _BODY_FONT
            ev.cell(row=er, column=4, value=cell.provenance.value).font = _BODY_FONT
            ev.cell(
                row=er, column=5,
                value=(round(cell.confidence, 2) if cell.confidence is not None else "")
            ).font = _BODY_FONT
            rc = ev.cell(row=er, column=6, value=cell.rationale)
            rc.font = _BODY_FONT
            rc.alignment = _WRAP
            evtxt = " | ".join(
                f"{e.doc}" + (f" p.{e.page}" if e.page else "") + (f": {e.quote}" if e.quote else "")
                for e in cell.evidence
            )
            ec = ev.cell(row=er, column=7, value=evtxt)
            ec.font = _BODY_FONT
            ec.alignment = _WRAP
            er += 1
    for col, w in zip("ABCDEFG", (18, 28, 8, 12, 10, 50, 60)):
        ev.column_dimensions[col].width = w

    # ---- Sheet 3: Scheme ------------------------------------------------ #
    sc = wb.create_sheet("Scheme")
    sc.cell(row=1, column=1, value="Project").font = _LABEL_FONT
    sc.cell(row=1, column=2, value=result.project_id).font = _BODY_FONT
    sc.cell(row=2, column=1, value="Method").font = _LABEL_FONT
    sc.cell(row=2, column=2, value=scheme.method.value).font = _BODY_FONT
    sc.cell(row=3, column=1, value="Scale").font = _LABEL_FONT
    sc.cell(
        row=3, column=2,
        value=f"{scheme.scale.min:g}-{scheme.scale.max:g} ({scheme.scale.type})"
    ).font = _BODY_FONT
    sc.cell(row=4, column=1, value="Cost handling").font = _LABEL_FONT
    sc.cell(row=4, column=2, value=scheme.cost_handling).font = _BODY_FONT
    sc.cell(row=5, column=1, value="Source").font = _LABEL_FONT
    sc.cell(row=5, column=2, value=scheme.source).font = _BODY_FONT
    if scheme.gates:
        sc.cell(row=6, column=1, value="Gates").font = _LABEL_FONT
        sc.cell(row=6, column=2, value="; ".join(scheme.gates)).font = _BODY_FONT
    sc.column_dimensions["A"].width = 16
    sc.column_dimensions["B"].width = 60

    wb.save(path)
    return path
